"""
Master Thesis — Jelle Vermeulen
Insider Ownership and Stock Returns: A Global FF5 Portfolio Analysis
Jan 2015 – Dec 2025  |  132 months  |  1,285 firms

HOW TO RUN
----------
1. Install packages if needed:
       pip install pandas numpy openpyxl statsmodels

2. Download FF5 factors from Ken French's data library:
       https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/data_library.html
   Download: "Developed Markets Factor and Returns" → monthly CSV
   Download: "Global MOM Factor"                    → monthly CSV (for FF6 robustness)
   Unzip both and set the paths in the CONFIG section below.

3. Set DATA_FILE to the path of your Excel file.

4. Run:  python thesis_analysis.py

OUTPUT FILES
------------
  results_global_EW.csv      — FF5 alphas, factor loadings, t-stats (equal-weighted)
  results_regional_EW.csv    — same, broken out by region (H4 / H5)
  portfolio_returns.csv      — raw monthly EW returns for all portfolios
  descriptive_table.csv      — Table 1: firm counts, ownership ranges, region splits
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import statsmodels.api as sm
import warnings
warnings.filterwarnings('ignore')


# ══════════════════════════════════════════════════════════════════════
# CONFIG — set your file paths here
# ══════════════════════════════════════════════════════════════════════

DATA_FILE = "Dataset master thesis Jelle Vermeulen.xlsx"

# Global factors (for H1, H2, H3)
FF5_GLOBAL  = "Developed_5_Factors.csv"

# Regional FF5 factors (for H4, H5)
FF5_NA      = "North_America_5_Factors.csv"
FF5_EU      = "Europe_5_Factors.csv"
FF5_APAC    = "Asia_Pacific_ex_Japan_5_Factors.csv"

# Momentum factors (for FF6 robustness — H3)
MOM_GLOBAL  = "Developed_MOM_Factor.csv"
MOM_NA      = "North_America_MOM_Factor.csv"
MOM_EU      = "Europe_MOM_Factor.csv"
MOM_APAC    = "Asia_Pacific_ex_Japan_MOM_Factor.csv"

# EM ex-Asia has no specific French factor file — global factors used as proxy
SAMPLE_START = "2015-01-01"
SAMPLE_END   = "2025-12-31"
NW_LAGS      = 6    # Newey-West lags (standard for monthly data)
SKIPROWS     = 6    # rows of text above column header in French CSV files


# ══════════════════════════════════════════════════════════════════════
# STEP 1 — LOAD AND CLEAN OWNERSHIP DATA
# ══════════════════════════════════════════════════════════════════════

print("Loading ownership data...")
wb      = load_workbook(DATA_FILE, read_only=True, data_only=True)
ws_own  = wb["Wood Owner firms "]
ws_ts   = wb["Time series data Returns"]
ws_mc   = wb["Time series data Market Cap"]
own_raw = list(ws_own.iter_rows(values_only=True))
ts_raw  = list(ws_ts.iter_rows(values_only=True))
mc_raw  = list(ws_mc.iter_rows(values_only=True))
wb.close()

own_df = pd.DataFrame(own_raw[1:], columns=own_raw[0])
own_df.rename(columns={"Returns tickers": "Returns_ticker"}, inplace=True)
own_df["Ownership"] = pd.to_numeric(own_df["Ownership"], errors="coerce")
own_df = own_df.dropna(subset=["Ownership", "Returns_ticker"])
own_df["Returns_ticker"] = own_df["Returns_ticker"].astype(str).str.strip()

# Exclude Rumble Inc. (ownership > 100% — data error)
own_df = own_df[own_df["Ownership"] <= 100]
own_df = own_df.drop_duplicates("Returns_ticker")

print(f"  Firms after cleaning: {len(own_df)}")


# ══════════════════════════════════════════════════════════════════════
# STEP 2 — ASSIGN OWNERSHIP QUINTILES
# ══════════════════════════════════════════════════════════════════════

own_df["Portfolio"] = pd.qcut(
    own_df["Ownership"], q=5, labels=["P1", "P2", "P3", "P4", "P5"]
)

REGION_MAP = {
    "United States and Canada":    "North America",
    "Europe":                      "Europe",
    "Asia / Pacific":              "Asia-Pacific",
    "Latin America and Caribbean": "EM ex-Asia",
    "Africa / Middle East":        "EM ex-Asia",
}
own_df["Region"] = own_df["Country"].map(REGION_MAP)

bp = own_df.groupby("Portfolio", observed=True)["Ownership"].agg(["min","max","count"])
print("\nQuintile breakpoints:")
print(bp.round(2).to_string())


# ══════════════════════════════════════════════════════════════════════
# STEP 3 — LOAD MONTHLY TOTAL RETURNS
# ══════════════════════════════════════════════════════════════════════

print("\nLoading return data...")
dates   = [r[0] for r in ts_raw[1:] if r[0] is not None]
matrix  = [r[1:] for r in ts_raw[1:] if r[0] is not None]
ts_cols = list(ts_raw[0][1:])

ret = pd.DataFrame(matrix, index=pd.to_datetime(dates), columns=ts_cols)
ret = ret.replace("#N/A", np.nan).apply(pd.to_numeric, errors="coerce")
ret = ret / 100                                    # % → decimal
ret = ret.loc[SAMPLE_START:SAMPLE_END]             # 132 months

# Drop tickers with zero data months (the 11 no-data tickers)
zero_data = ret.columns[ret.notna().sum() == 0]
ret = ret.drop(columns=zero_data)
print(f"  Dropped {len(zero_data)} zero-data tickers")
print(f"  Return matrix: {ret.shape[0]} months × {ret.shape[1]} firms")


# ══════════════════════════════════════════════════════════════════════
# STEP 4 — BUILD EQUAL-WEIGHTED PORTFOLIO RETURNS
# ══════════════════════════════════════════════════════════════════════

matched    = own_df[own_df["Returns_ticker"].isin(ret.columns)].copy()
fid_port   = matched.set_index("Returns_ticker")["Portfolio"].to_dict()
fid_region = matched.set_index("Returns_ticker")["Region"].to_dict()
print(f"  Matched firms: {len(matched)}")

PORTFOLIOS = ["P1", "P2", "P3", "P4", "P5"]
REGIONS    = ["North America", "Europe", "Asia-Pacific", "EM ex-Asia"]

def winsorise(s, lo=0.01, hi=0.99):
    """Winsorise a return series at 1st and 99th percentile."""
    return s.clip(s.quantile(lo), s.quantile(hi))

def build_ew_portfolios(ret_df, port_map):
    """
    Build equal-weighted portfolios P1–P5 and HML spread.
    Returns a DataFrame of monthly returns (decimal).
    """
    ew = pd.DataFrame(index=ret_df.index)
    for p in PORTFOLIOS:
        cols = [t for t, port in port_map.items() if port == p and t in ret_df.columns]
        if len(cols) > 0:
            ew[p] = winsorise(ret_df[cols].mean(axis=1, skipna=True))
        else:
            ew[p] = np.nan
    ew["HML"] = ew["P5"] - ew["P1"]
    return ew

# Global portfolios
ew_global = build_ew_portfolios(ret, fid_port)

# Regional portfolios
ew_regional = {}
for reg in REGIONS:
    reg_map = {t: p for t, p in fid_port.items() if fid_region.get(t) == reg}
    ew_regional[reg] = build_ew_portfolios(ret, reg_map)

print("\nGlobal EW portfolio raw returns (% per month):")
print((ew_global.mean() * 100).round(3).to_string())

# ── Value-weighted portfolios ──────────────────────────────────────
print("\nBuilding value-weighted portfolios...")

mc_dates  = [r[0] for r in mc_raw[1:] if r[0] is not None]
mc_matrix = [r[1:] for r in mc_raw[1:] if r[0] is not None]
mc_cols   = list(mc_raw[0][1:])

mktcap = pd.DataFrame(mc_matrix, index=pd.to_datetime(mc_dates), columns=mc_cols)
mktcap = mktcap.replace("#N/A", np.nan).apply(pd.to_numeric, errors="coerce")
mktcap = mktcap.loc[SAMPLE_START:SAMPLE_END]

# Align columns with return data
common_cols = [c for c in ret.columns if c in mktcap.columns]
mktcap = mktcap[common_cols]

def build_vw_portfolios(ret_df, mc_df, port_map):
    """Value-weighted portfolios using lagged market cap as weights."""
    labels = ["P1","P2","P3","P4","P5"]
    vw = pd.DataFrame(index=ret_df.index)
    for p in labels:
        cols = [t for t, port in port_map.items() if port == p
                and t in ret_df.columns and t in mc_df.columns]
        port_ret = ret_df[cols]
        port_mc  = mc_df[cols].shift(1)      # lagged market cap as weight
        weights  = port_mc.div(port_mc.sum(axis=1), axis=0)
        raw = (port_ret * weights).sum(axis=1, skipna=True)
        vw[p] = raw.clip(raw.quantile(0.01), raw.quantile(0.99))
    vw["HML"] = vw["P5"] - vw["P1"]
    return vw

vw_global = build_vw_portfolios(ret, mktcap, fid_port)
print(f"  VW global portfolios built")
print(f"  Raw VW monthly returns (%):")
print((vw_global.mean() * 100).round(3).to_string())


# ══════════════════════════════════════════════════════════════════════
# STEP 5 — LOAD FAMA-FRENCH FACTORS
# ══════════════════════════════════════════════════════════════════════
#
# French CSV format:
#   Rows 0–5  : descriptive text (skipped via SKIPROWS=6)
#   Row 6     : column header  e.g. ",Mkt-RF,SMB,HML,RMW,CMA,RF"
#   Row 7+    : data           e.g. "201501,3.42,0.21,..."
#   Date format: YYYYMM  |  Values in %  |  Missing = -99.99
#   Momentum column: WML (renamed to UMD internally)

def load_french_factors(filepath, start=SAMPLE_START, end=SAMPLE_END):
    """
    Load a Ken French CSV factor file and return a cleaned DataFrame.
    Returns None if the file is not found (non-fatal for regional files).
    French files contain monthly data followed by an annual section —
    both are handled by converting to numeric and dropping non-numeric rows.
    """
    try:
        raw = pd.read_csv(filepath, skiprows=SKIPROWS, header=0,
                          index_col=0, na_values=["-99.99", ""])
        raw.index = raw.index.astype(str).str.strip()

        # Keep only rows where the index is a 6-digit YYYYMM date
        # This drops the annual section at the bottom of French files
        raw = raw[raw.index.str.match(r"^\d{6}$")]
        raw.index = pd.to_datetime(raw.index, format="%Y%m", errors="coerce")
        raw.index = raw.index + pd.offsets.MonthEnd(0)  # align to month-end like portfolio returns
        raw = raw[raw.index.notna()]

        # Convert all columns to numeric — drops any remaining text rows
        raw.columns = [c.strip() for c in raw.columns]
        raw = raw.apply(pd.to_numeric, errors="coerce")
        raw = raw.dropna(how="all")
        raw = raw / 100                       # % → decimal
        raw = raw.loc[start:end]
        if "WML" in raw.columns:
            raw = raw.rename(columns={"WML": "UMD"})
        print(f"  Loaded: {filepath}  ({len(raw)} months, cols: {list(raw.columns)})")
        return raw
    except FileNotFoundError:
        print(f"  NOT FOUND: {filepath}  — set correct filename in CONFIG")
        return None

print("\nLoading Fama-French factors...")

# Global
ff5_global = load_french_factors(FF5_GLOBAL)
mom_global = load_french_factors(MOM_GLOBAL)

if ff5_global is None:
    print("\n*** Global FF5 file missing — cannot run regressions. ***")
    print("Set FF5_GLOBAL in the CONFIG section to the correct filename.")
    ew_global.to_csv("portfolio_returns.csv")
    exit()

# Build FF6 global (FF5 + momentum)
ff6_global = None
if mom_global is not None and "UMD" in mom_global.columns:
    ff6_global = ff5_global.join(mom_global[["UMD"]], how="left")

# Regional
ff5_regional = {
    "North America": load_french_factors(FF5_NA),
    "Europe":        load_french_factors(FF5_EU),
    "Asia-Pacific":  load_french_factors(FF5_APAC),
    "EM ex-Asia":    ff5_global,   # proxy: global factors (noted as limitation)
}

mom_regional = {
    "North America": load_french_factors(MOM_NA),
    "Europe":        load_french_factors(MOM_EU),
    "Asia-Pacific":  load_french_factors(MOM_APAC),
    "EM ex-Asia":    mom_global,   # proxy
}

# Build FF6 per region
ff6_regional = {}
for reg in REGIONS:
    f5  = ff5_regional[reg]
    mom = mom_regional[reg]
    if f5 is not None and mom is not None and "UMD" in mom.columns:
        ff6_regional[reg] = f5.join(mom[["UMD"]], how="left")
    else:
        ff6_regional[reg] = None


# ══════════════════════════════════════════════════════════════════════
# STEP 6 — FF5 REGRESSION FUNCTION
# ══════════════════════════════════════════════════════════════════════

FACTOR_COLS_FF5 = ["Mkt-RF", "SMB", "HML", "RMW", "CMA"]
FACTOR_COLS_FF6 = ["Mkt-RF", "SMB", "HML", "RMW", "CMA", "UMD"]

def run_ff_regression(portfolio_returns, factors, factor_cols, rf_col="RF", nw_lags=NW_LAGS):
    """
    Run FF5 (or FF6) time-series regression for each portfolio.
    Returns a DataFrame with alpha, t-stat, factor betas, adj R2.

    Specification:
        R_p,t - RF_t = alpha + B1*MKT + B2*SMB + B3*HML + B4*RMW + B5*CMA + e_t
    Standard errors: Newey-West with nw_lags lags.
    """
    results = []
    rf = factors[rf_col]

    for col in portfolio_returns.columns:
        y = portfolio_returns[col].dropna()

        # For HML portfolio, no RF subtraction (it's already a spread)
        if col == "HML":
            y_excess = y
        else:
            y_excess = (y - rf).dropna()

        # Align factors to available return dates
        X = factors[factor_cols].reindex(y_excess.index).dropna()
        y_excess = y_excess.reindex(X.index)

        if len(y_excess) < 24:
            results.append({"Portfolio": col, "Note": "Insufficient data"})
            continue

        X_const = sm.add_constant(X)
        model   = sm.OLS(y_excess, X_const).fit(
            cov_type="HAC",
            cov_kwds={"maxlags": nw_lags}
        )

        row = {
            "Portfolio":  col,
            "Alpha (% pm)":   round(model.params["const"] * 100, 4),
            "Alpha (% pa)":   round(model.params["const"] * 12 * 100, 3),
            "t-stat (NW)":    round(model.tvalues["const"], 3),
            "p-value":        round(model.pvalues["const"], 4),
            "Adj R²":         round(model.rsquared_adj, 4),
            "N months":       int(model.nobs),
        }
        for f in factor_cols:
            row[f"β {f}"] = round(model.params[f], 4)

        results.append(row)

    return pd.DataFrame(results).set_index("Portfolio")


# ══════════════════════════════════════════════════════════════════════
# STEP 7 — RUN GLOBAL REGRESSIONS (H1, H2, H3)
# ══════════════════════════════════════════════════════════════════════

print("\n" + "="*65)
print("FF5 REGRESSION — GLOBAL EW PORTFOLIOS  (H1 and H2)")
print("="*65)

res_global_ff5 = run_ff_regression(ew_global, ff5_global, FACTOR_COLS_FF5)
print(res_global_ff5[["Alpha (% pm)", "Alpha (% pa)", "t-stat (NW)", "p-value", "Adj R²"]].to_string())

if ff6_global is not None:
    print("\n" + "="*65)
    print("FF6 REGRESSION — GLOBAL EW PORTFOLIOS  (H3 robustness)")
    print("="*65)
    res_global_ff6 = run_ff_regression(ew_global, ff6_global, FACTOR_COLS_FF6)
    print(res_global_ff6[["Alpha (% pm)", "Alpha (% pa)", "t-stat (NW)", "p-value", "Adj R²"]].to_string())

# Value-weighted FF5 regression (Table R3)
print("\n" + "="*65)
print("FF5 REGRESSION — GLOBAL VW PORTFOLIOS  (robustness)")
print("="*65)
res_global_vw = run_ff_regression(vw_global, ff5_global, FACTOR_COLS_FF5)
print(res_global_vw[["Alpha (% pm)", "Alpha (% pa)", "t-stat (NW)", "p-value", "Adj R²"]].to_string())


# ══════════════════════════════════════════════════════════════════════
# STEP 8 — RUN REGIONAL REGRESSIONS (H4 and H5)
# ══════════════════════════════════════════════════════════════════════

print("\n" + "="*65)
print("FF5 REGRESSION — REGIONAL EW PORTFOLIOS  (H4 and H5)")
print("="*65)

regional_results    = {}
regional_results_ff6 = {}

for reg in REGIONS:
    factors_reg = ff5_regional[reg]
    if factors_reg is None:
        print(f"\n  {reg}: skipped — factor file missing")
        continue
    print(f"\n  Region: {reg}")
    res = run_ff_regression(ew_regional[reg], factors_reg, FACTOR_COLS_FF5)
    regional_results[reg] = res
    print(res[["Alpha (% pm)", "Alpha (% pa)", "t-stat (NW)", "p-value", "Adj R²"]].to_string())

    # FF6 robustness per region
    if ff6_regional.get(reg) is not None:
        res6 = run_ff_regression(ew_regional[reg], ff6_regional[reg], FACTOR_COLS_FF6)
        regional_results_ff6[reg] = res6


# ══════════════════════════════════════════════════════════════════════
# STEP 8B — ROBUSTNESS REGRESSIONS
# Split-sample and alternative sorts — run here so all regressions
# are in one script and thesis_output.py only reads result files.
# ══════════════════════════════════════════════════════════════════════

print("\nRunning robustness regressions...")

# Rebuild full quintile EW returns for robustness (same as ew_global)
ew_q5 = ew_global.copy()

# ── Split-sample ───────────────────────────────────────────────────
SPLIT    = "2019-12-31"
pre_ret  = ew_q5.loc[:SPLIT]
post_ret = ew_q5.loc[pd.Timestamp(SPLIT) + pd.offsets.MonthBegin(1):]
ff5_pre  = ff5_global.loc[:SPLIT]
ff5_post = ff5_global.loc[pd.Timestamp(SPLIT) + pd.offsets.MonthBegin(1):]

res_pre  = run_ff_regression(pre_ret,  ff5_pre,  FACTOR_COLS_FF5)
res_post = run_ff_regression(post_ret, ff5_post, FACTOR_COLS_FF5)

print(f"  Pre-COVID ({len(pre_ret)} months): HML alpha = "
      f"{res_pre.loc['HML','Alpha (% pm)']:.4f}%  "
      f"(t={res_pre.loc['HML','t-stat (NW)']:.3f})")
print(f"  Post-COVID ({len(post_ret)} months): HML alpha = "
      f"{res_post.loc['HML','Alpha (% pm)']:.4f}%  "
      f"(t={res_post.loc['HML','t-stat (NW)']:.3f})")

# ── Alternative sorts ──────────────────────────────────────────────
ownership_series = matched.set_index("Returns_ticker")["Ownership"]

def build_ew_custom(ret_df, port_map):
    labels = sorted(set(port_map.values()))
    out = pd.DataFrame(index=ret_df.index)
    for lbl in labels:
        cols = [t for t, p in port_map.items() if p == lbl and t in ret_df.columns]
        raw  = ret_df[cols].mean(axis=1, skipna=True)
        out[lbl] = raw.clip(raw.quantile(0.01), raw.quantile(0.99))
    return out

# Tercile sort
t1_bp = ownership_series.quantile(1/3)
t2_bp = ownership_series.quantile(2/3)
map_t3 = {t: ("T1" if v < t1_bp else "T2" if v < t2_bp else "T3")
           for t, v in ownership_series.items()}
ew_t3 = build_ew_custom(ret, map_t3)
ew_t3["HML"] = ew_t3["T3"] - ew_t3["T1"]
res_t3 = run_ff_regression(ew_t3, ff5_global, FACTOR_COLS_FF5)

# Decile sort
dec_breaks = [ownership_series.quantile(i/10) for i in range(11)]
def assign_decile(v):
    for i in range(9):
        if v < dec_breaks[i+1]:
            return f"D{i+1}"
    return "D10"
map_d10 = {t: assign_decile(v) for t, v in ownership_series.items()}
ew_d10  = build_ew_custom(ret, map_d10)
ew_d10["HML"] = ew_d10["D10"] - ew_d10["D1"]
res_d10 = run_ff_regression(ew_d10, ff5_global, FACTOR_COLS_FF5)

print(f"  Tercile HML alpha = "
      f"{res_t3.loc['HML','Alpha (% pm)']:.4f}%  "
      f"(t={res_t3.loc['HML','t-stat (NW)']:.3f})")
print(f"  Decile HML alpha  = "
      f"{res_d10.loc['HML','Alpha (% pm)']:.4f}%  "
      f"(t={res_d10.loc['HML','t-stat (NW)']:.3f})")


# ══════════════════════════════════════════════════════════════════════
# STEP 9 — SAVE ALL RESULTS
# ══════════════════════════════════════════════════════════════════════

# Global results
ew_global.to_csv("portfolio_returns.csv")
res_global_ff5.to_csv("results_global_FF5_EW.csv")
if ff6_global is not None:
    res_global_ff6.to_csv("results_global_FF6_EW.csv")
res_global_vw.to_csv("results_global_FF5_VW.csv")

# Regional FF5 results — one sheet per region
with pd.ExcelWriter("results_regional_FF5_EW.xlsx", engine="openpyxl") as w:
    for reg, res in regional_results.items():
        sheet = reg.replace(" ","_").replace("-","_")[:30]
        res.to_excel(w, sheet_name=sheet)

# Regional FF6 results
if regional_results_ff6:
    with pd.ExcelWriter("results_regional_FF6_EW.xlsx", engine="openpyxl") as w:
        for reg, res in regional_results_ff6.items():
            sheet = reg.replace(" ","_").replace("-","_")[:30]
            res.to_excel(w, sheet_name=sheet)

# Robustness results
with pd.ExcelWriter("results_splitsample_FF5_EW.xlsx", engine="openpyxl") as w:
    res_pre.to_excel(w,  sheet_name="Pre_COVID_2015_2019")
    res_post.to_excel(w, sheet_name="Post_COVID_2020_2025")

with pd.ExcelWriter("results_alternative_sorts_FF5.xlsx", engine="openpyxl") as w:
    res_t3.to_excel(w,  sheet_name="Tercile_Sort")
    res_d10.to_excel(w, sheet_name="Decile_Sort")

# Descriptive table
desc = matched.groupby(["Portfolio","Region"], observed=True).agg(
    N_firms        = ("Returns_ticker", "count"),
    Ownership_mean = ("Ownership", "mean"),
    Ownership_min  = ("Ownership", "min"),
    Ownership_max  = ("Ownership", "max"),
).round(2)
desc.to_csv("descriptive_table.csv")

# Firm-level quintile assignments
matched[["FirmID","Company","Country","Region","Ownership","Portfolio","Returns_ticker"]]\
    .to_csv("firms_quintiles.csv", index=False)

print("\n" + "="*65)
print("DONE — files saved:")
print("  portfolio_returns.csv              — monthly EW returns P1–P5 + HML")
print("  results_global_FF5_EW.csv          — global FF5 alphas EW (H1, H2)")
print("  results_global_FF6_EW.csv          — global FF6 alphas EW (H3)")
print("  results_global_FF5_VW.csv          — global FF5 alphas VW (R3)")
print("  results_regional_FF5_EW.xlsx       — regional FF5 alphas (H4, H5)")
print("  results_regional_FF6_EW.xlsx       — regional FF6 alphas (H4, H5)")
print("  results_splitsample_FF5_EW.xlsx    — split-sample robustness (R1)")
print("  results_alternative_sorts_FF5.xlsx — alternative sorts robustness (R2)")
print("  descriptive_table.csv              — firm counts and ownership ranges")
print("  firms_quintiles.csv                — 1285 firms with quintile labels")
print("\nNow run:  python thesis_output.py")
print("="*65)
