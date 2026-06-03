"""
Master Thesis — Jelle Vermeulen
Output Script: All Tables and All Figures

Run AFTER thesis_analysis.py has produced all result files.

Usage:     python thesis_output.py
Produces:  thesis_tables.xlsx   — all 7 academic tables (Tables 1-5, R1, R2)
           figure1_alpha_quintiles.pdf/png
           figure2_cumulative_returns.pdf/png
           figure3_regional_alphas.pdf/png
           figure4_factor_loadings.pdf/png
           figure5_ff5_vs_ff6.pdf/png
           figure6_decile_alphas.pdf/png
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.dates as mdates
import matplotlib.cm as cm
import matplotlib.colors as mcolors
from matplotlib.colors import TwoSlopeNorm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import warnings
warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════
# LOAD ALL RESULT FILES
# ══════════════════════════════════════════════════════════════════════

print("Loading result files...")

desc   = pd.read_csv("descriptive_table.csv")
firms  = pd.read_csv("firms_quintiles.csv")
ret    = pd.read_csv("portfolio_returns.csv",       index_col=0, parse_dates=True)
ff5    = pd.read_csv("results_global_FF5_EW.csv",   index_col=0)
ff6    = pd.read_csv("results_global_FF6_EW.csv",   index_col=0)
vw     = pd.read_csv("results_global_FF5_VW.csv",   index_col=0)
sp_xl  = pd.ExcelFile("results_splitsample_FF5_EW.xlsx")
alt_xl = pd.ExcelFile("results_alternative_sorts_FF5.xlsx")
reg_xl = pd.ExcelFile("results_regional_FF5_EW.xlsx")

res_pre  = sp_xl.parse("Pre_COVID_2015_2019",  index_col=0)
res_post = sp_xl.parse("Post_COVID_2020_2025", index_col=0)
res_t3   = alt_xl.parse("Tercile_Sort",        index_col=0)
res_d10  = alt_xl.parse("Decile_Sort",         index_col=0)

reg_sheets = {}
sheet_to_region = {
    "North_America": "North America",
    "Europe":        "Europe",
    "Asia_Pacific":  "Asia-Pacific",
    "EM_ex_Asia":    "EM ex-Asia",
}
for sheet in reg_xl.sheet_names:
    reg = sheet_to_region.get(sheet, sheet)
    reg_sheets[reg] = reg_xl.parse(sheet, index_col=0)

PORTFOLIOS = ["P1","P2","P3","P4","P5","HML"]
REGIONS    = ["North America","Europe","Asia-Pacific","EM ex-Asia"]
DECILES    = [f"D{i}" for i in range(1,11)]

# Ownership helper values
own = firms["Ownership"]
t1_break = own.quantile(1/3)
t2_break = own.quantile(2/3)
cnt_t3 = {
    "T1": int((own < t1_break).sum()),
    "T2": int(((own >= t1_break)&(own < t2_break)).sum()),
    "T3": int((own >= t2_break).sum()),
}
dec_ranges = {
    "D1":"20.0–23.4%","D2":"23.4–26.4%","D3":"26.5–29.6%","D4":"29.6–34.0%",
    "D5":"34.0–38.8%","D6":"38.8–43.6%","D7":"43.7–49.9%","D8":"49.9–56.7%",
    "D9":"56.8–68.3%","D10":"68.4–99.8%",
}

# d10_ranges: numeric tuples (min%, max%) per decile — used in build_r2 table headers
_dec_breaks = [own.quantile(i/10) for i in range(11)]
d10_ranges  = {}
for _i in range(10):
    _lo = _dec_breaks[_i]
    _hi = _dec_breaks[_i+1]
    _mask = (own >= _lo) & (own < _hi) if _i < 9 else (own >= _lo)
    _vals = own[_mask]
    if len(_vals) > 0:
        d10_ranges[f"D{_i+1}"] = (float(_vals.min()), float(_vals.max()))

print("  All files loaded.\n")


# ══════════════════════════════════════════════════════════════════════
# SHARED HELPERS
# ══════════════════════════════════════════════════════════════════════

def stars(pv):
    return "***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""

def fmt(val, decimals=4):
    if pd.isna(val): return ""
    return f"{val:.{decimals}f}"

def fmt_pct(val, decimals=2):
    if pd.isna(val): return ""
    return f"{val:.{decimals}f}%"

def save_fig(fig, name):
    fig.savefig(f"{name}.pdf")
    fig.savefig(f"{name}.png")
    print(f"  Saved: {name}.pdf / .png")




# ══════════════════════════════════════════════════════════════════════
# TABLES — Styles, builders, and workbook assembly
# ══════════════════════════════════════════════════════════════════════

# ── Excel style constants ──────────────────────────────────────────────
FONT_BODY   = Font(name="Times New Roman", size=10)
FONT_BOLD   = Font(name="Times New Roman", size=10, bold=True)
FONT_TITLE  = Font(name="Times New Roman", size=11, bold=True)
FONT_NOTE   = Font(name="Times New Roman", size=9, italic=True)
FONT_PANEL  = Font(name="Times New Roman", size=10, bold=True, italic=True)

ALIGN_L  = Alignment(horizontal="left",  vertical="center", wrap_text=False)
ALIGN_R  = Alignment(horizontal="right", vertical="center", wrap_text=False)
ALIGN_C  = Alignment(horizontal="center",vertical="center", wrap_text=False)

THICK = Side(style="medium", color="000000")
THIN  = Side(style="thin",   color="000000")
NONE  = Side(style=None)

def top_border():    return Border(top=THICK, bottom=NONE, left=NONE, right=NONE)
def header_border(): return Border(top=NONE,  bottom=THIN,  left=NONE, right=NONE)
def bottom_border(): return Border(top=THICK, bottom=NONE, left=NONE, right=NONE)
def no_border():     return Border(top=NONE,  bottom=NONE, left=NONE, right=NONE)

def style_row(ws, row_num, col_start, col_end, font=None, align=None, border=None):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        if font:   cell.font   = font
        if align:  cell.alignment = align
        if border: cell.border = border

def write_cell(ws, row, col, value, font=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = font   or FONT_BODY
    cell.alignment = align or ALIGN_R
    if border: cell.border = border
    return cell

# Short alias used in robustness table builders
def wc(ws, row, col, val, font=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = font  or FONT_BODY
    cell.alignment = align or ALIGN_C
    if border: cell.border = border
    return cell


# ══════════════════════════════════════════════════════════════════════
# TABLE 1 — SAMPLE CHARACTERISTICS
# Follows: Wood (2025) Table 2 and Fahlenbrach (2009) Table 1
# ══════════════════════════════════════════════════════════════════════

def build_table1(wb):
    ws = wb.create_sheet("Table 1")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 14

    r = 1
    # Title
    ws.cell(r,1,"Table 1").font = FONT_TITLE
    ws.cell(r,1).alignment = ALIGN_L
    r += 1
    ws.cell(r,1,"Sample Characteristics: Insider Ownership Quintile Portfolios").font = FONT_TITLE
    ws.cell(r,1).alignment = ALIGN_L
    r += 2

    # ── Panel A: Firm distribution ──────────────────────────────────
    ws.cell(r,1,"Panel A: Firm Distribution by Ownership Quintile and Region").font = FONT_PANEL
    ws.cell(r,1).alignment = ALIGN_L
    r += 1

    # Column headers
    hdrs = ["Portfolio","Ownership Range","N Firms","Mean Own. (%)","Asia-Pacific","Europe","N. America","EM ex-Asia"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(r, i, h)
        c.font = FONT_BOLD
        c.alignment = ALIGN_C if i > 1 else ALIGN_L
        c.border = header_border()
    style_row(ws, r, 1, len(hdrs), border=Border(top=THICK, bottom=THIN))
    r += 1

    # Ownership ranges from descriptive table
    ranges = {
        "P1": "20.00 – 26.46%",
        "P2": "26.48 – 34.04%",
        "P3": "34.05 – 43.71%",
        "P4": "43.74 – 56.84%",
        "P5": "56.87 – 99.81%",
    }
    firm_ct = firms.groupby(["Portfolio","Region"])["FirmID"].count().unstack(fill_value=0)
    region_order = ["Asia-Pacific","Europe","North America","EM ex-Asia"]
    port_labels = {"P1":"P1 (Low)","P2":"P2","P3":"P3","P4":"P4","P5":"P5 (High)"}

    totals = {"N":0, "APAC":0, "EU":0, "NA":0, "EM":0}
    mean_own = firms.groupby("Portfolio")["Ownership"].mean()

    for p in ["P1","P2","P3","P4","P5"]:
        row_data = firm_ct.loc[p] if p in firm_ct.index else pd.Series(0, index=region_order)
        n   = row_data.sum()
        ap  = row_data.get("Asia-Pacific", 0)
        eu  = row_data.get("Europe", 0)
        na  = row_data.get("North America", 0)
        em  = row_data.get("EM ex-Asia", 0)
        mo  = mean_own.get(p, np.nan)
        totals["N"]  += n; totals["APAC"] += ap; totals["EU"] += eu
        totals["NA"] += na; totals["EM"]  += em

        write_cell(ws, r, 1, port_labels[p], FONT_BODY, ALIGN_L)
        write_cell(ws, r, 2, ranges[p],      FONT_BODY, ALIGN_C)
        write_cell(ws, r, 3, n,              FONT_BODY, ALIGN_C)
        write_cell(ws, r, 4, f"{mo:.2f}",   FONT_BODY, ALIGN_C)
        write_cell(ws, r, 5, ap,            FONT_BODY, ALIGN_C)
        write_cell(ws, r, 6, eu,            FONT_BODY, ALIGN_C)
        write_cell(ws, r, 7, na,            FONT_BODY, ALIGN_C)
        write_cell(ws, r, 8, em,            FONT_BODY, ALIGN_C)
        r += 1

    # Total row
    style_row(ws, r, 1, 8, border=Border(top=THIN))
    write_cell(ws, r, 1, "Total", FONT_BOLD, ALIGN_L, Border(top=THIN))
    write_cell(ws, r, 2, "",      FONT_BOLD, ALIGN_C, Border(top=THIN))
    write_cell(ws, r, 3, totals["N"],    FONT_BOLD, ALIGN_C, Border(top=THIN))
    mo_all = firms["Ownership"].mean()
    write_cell(ws, r, 4, f"{mo_all:.2f}", FONT_BOLD, ALIGN_C, Border(top=THIN))
    write_cell(ws, r, 5, totals["APAC"], FONT_BOLD, ALIGN_C, Border(top=THIN))
    write_cell(ws, r, 6, totals["EU"],   FONT_BOLD, ALIGN_C, Border(top=THIN))
    write_cell(ws, r, 7, totals["NA"],   FONT_BOLD, ALIGN_C, Border(top=THIN))
    write_cell(ws, r, 8, totals["EM"],   FONT_BOLD, ALIGN_C, Border(top=THIN))
    style_row(ws, r, 1, 8, border=Border(top=THIN, bottom=THICK))
    r += 3

    # ── Panel B: Portfolio return summary statistics ─────────────────
    ws.cell(r,1,"Panel B: Portfolio Return Summary Statistics").font = FONT_PANEL
    ws.cell(r,1).alignment = ALIGN_L
    r += 1

    col_labels = ["P1\n(Low)","P2","P3","P4","P5\n(High)","HML\n(P5−P1)"]
    write_cell(ws, r, 1, "",  FONT_BOLD, ALIGN_L, Border(top=THICK, bottom=THIN))
    for i, lbl in enumerate(col_labels, 2):
        c = write_cell(ws, r, i, lbl, FONT_BOLD, ALIGN_C, Border(top=THICK, bottom=THIN))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1

    stats_rows = [
        ("Mean return (% per month)",  lambda p: ret[p].mean()*100,   "%.3f"),
        ("Mean return (% per year)",   lambda p: ret[p].mean()*1200,  "%.2f"),
        ("Std. dev. (% per month)",    lambda p: ret[p].std()*100,    "%.3f"),
        ("Minimum (% per month)",      lambda p: ret[p].min()*100,    "%.3f"),
        ("Maximum (% per month)",      lambda p: ret[p].max()*100,    "%.3f"),
        ("Sharpe ratio (annualised)",  lambda p: ret[p].mean()/ret[p].std()*np.sqrt(12), "%.3f"),
        ("Observations (months)",      lambda p: ret[p].notna().sum(), "%d"),
    ]

    for label, fn, fmt_str in stats_rows:
        write_cell(ws, r, 1, label, FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTFOLIOS, 2):
            val = fn(p)
            write_cell(ws, r, i, fmt_str % val, FONT_BODY, ALIGN_C)
        r += 1

    style_row(ws, r-1, 1, 7, border=Border(bottom=THICK))

    # Notes
    r += 1
    note = (
        "Notes: This table reports characteristics of the five ownership-sorted "
        "portfolios and the high-minus-low (HML) spread portfolio. Panel A shows the "
        "distribution of 1,285 firms from Wood (2025) across ownership quintiles and "
        "regions. Ownership ranges are determined by equal-size quintile sorts on "
        "insider ownership as of 2025. Sample period: January 2015 to December 2025 "
        "(132 months). Returns are equal weighted, in USD, and winsorised at the "
        "1st and 99th percentile. Sharpe ratio is annualised: (mean monthly return / "
        "monthly std dev) × √12."
    )
    ws.cell(r, 1, note).font = FONT_NOTE
    ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=8)


# ══════════════════════════════════════════════════════════════════════
# TABLE 2 — FF5 GLOBAL RESULTS  (Main result, H1 and H2)
# Follows: Fahlenbrach (2009, JFQA) Table 6 format
# ══════════════════════════════════════════════════════════════════════

def build_regression_table(wb, sheet_name, title, subtitle, data, factor_cols, note_text):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    for col in ["B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 13

    r = 1
    ws.cell(r, 1, title).font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_L
    r += 1
    ws.cell(r, 1, subtitle).font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_L
    r += 2

    # Column headers
    col_labels = ["P1\n(Low)", "P2", "P3", "P4", "P5\n(High)", "HML\n(P5−P1)"]
    write_cell(ws, r, 1, "", FONT_BOLD, ALIGN_L, Border(top=THICK, bottom=THIN))
    for i, lbl in enumerate(col_labels, 2):
        c = write_cell(ws, r, i, lbl, FONT_BOLD, ALIGN_C, Border(top=THICK, bottom=THIN))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1

    # ── Alpha row (coefficient + stars) ───────────────────────────
    write_cell(ws, r, 1, "α (% per month)", FONT_BOLD, ALIGN_L)
    for i, p in enumerate(PORTFOLIOS, 2):
        a  = data.loc[p, "Alpha (% pm)"]
        pv = data.loc[p, "p-value"]
        write_cell(ws, r, i, f"{a:.4f}{stars(pv)}", FONT_BOLD, ALIGN_C)
    r += 1

    # T-stat row (in parentheses, indented label)
    write_cell(ws, r, 1, "  [t-statistic]", FONT_BODY, ALIGN_L)
    for i, p in enumerate(PORTFOLIOS, 2):
        t = data.loc[p, "t-stat (NW)"]
        write_cell(ws, r, i, f"({t:.3f})", FONT_BODY, ALIGN_C)
    r += 1

    # Alpha annualised
    write_cell(ws, r, 1, "α (% per year)", FONT_BODY, ALIGN_L)
    for i, p in enumerate(PORTFOLIOS, 2):
        pa = data.loc[p, "Alpha (% pa)"]
        write_cell(ws, r, i, f"{pa:.3f}", FONT_BODY, ALIGN_C)
    r += 2

    # ── Factor loadings ────────────────────────────────────────────
    factor_labels = {
        "β Mkt-RF": "β  MKT-RF",
        "β SMB":    "β  SMB",
        "β HML":    "β  HML",
        "β RMW":    "β  RMW",
        "β CMA":    "β  CMA",
        "β UMD":    "β  UMD (Momentum)",
    }
    for fc in factor_cols:
        lbl = factor_labels.get(fc, fc)
        write_cell(ws, r, 1, lbl, FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTFOLIOS, 2):
            val = data.loc[p, fc] if fc in data.columns else np.nan
            write_cell(ws, r, i, fmt(val, 4), FONT_BODY, ALIGN_C)
        r += 1

    r += 1  # blank row before summary stats

    # ── Adj R² ────────────────────────────────────────────────────
    write_cell(ws, r, 1, "Adj. R²", FONT_BODY, ALIGN_L)
    for i, p in enumerate(PORTFOLIOS, 2):
        write_cell(ws, r, i, fmt(data.loc[p, "Adj R²"], 4), FONT_BODY, ALIGN_C)
    r += 1

    # Observations
    write_cell(ws, r, 1, "Observations (months)", FONT_BODY, ALIGN_L)
    for i, p in enumerate(PORTFOLIOS, 2):
        write_cell(ws, r, i, int(data.loc[p, "N months"]), FONT_BODY, ALIGN_C)
    r += 1

    # Bottom border
    style_row(ws, r-1, 1, 7, border=Border(bottom=THICK))

    # Notes
    r += 1
    ws.cell(r, 1, note_text).font = FONT_NOTE
    ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=7)


# ══════════════════════════════════════════════════════════════════════
# TABLE 4 — REGIONAL RESULTS  (H4 and H5)
# Panel structure: one panel per region
# ══════════════════════════════════════════════════════════════════════

def build_table4(wb, reg_data, title, subtitle, note_text):
    ws = wb.create_sheet("Table 4 Regional")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    for col in ["B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 13

    r = 1
    ws.cell(r, 1, title).font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_L
    r += 1
    ws.cell(r, 1, subtitle).font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_L
    r += 2

    firm_counts = firms.groupby("Region")["FirmID"].count()
    panel_labels = {
        "North America": "Panel A",
        "Europe":        "Panel B",
        "Asia-Pacific":  "Panel C",
        "EM ex-Asia":    "Panel D",
    }
    col_labels = ["P1\n(Low)", "P2", "P3", "P4", "P5\n(High)", "HML\n(P5−P1)"]

    for panel_idx, region in enumerate(REGIONS):
        data = reg_data[region]
        n_firms = firm_counts.get(region, 0)
        panel_lbl = panel_labels[region]

        # Panel header
        panel_title = f"{panel_lbl}: {region}  (N = {n_firms} firms)"
        ws.cell(r, 1, panel_title).font = FONT_PANEL
        ws.cell(r, 1).alignment = ALIGN_L
        r += 1

        # Column headers (only on first panel and repeat)
        write_cell(ws, r, 1, "", FONT_BOLD, ALIGN_L, Border(top=THICK, bottom=THIN))
        for i, lbl in enumerate(col_labels, 2):
            c = write_cell(ws, r, i, lbl, FONT_BOLD, ALIGN_C, Border(top=THICK, bottom=THIN))
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1

        # Alpha
        write_cell(ws, r, 1, "α (% per month)", FONT_BOLD, ALIGN_L)
        for i, p in enumerate(PORTFOLIOS, 2):
            if p in data.index:
                a  = data.loc[p, "Alpha (% pm)"]
                pv = data.loc[p, "p-value"]
                write_cell(ws, r, i, f"{a:.4f}{stars(pv)}", FONT_BOLD, ALIGN_C)
        r += 1

        # T-stat
        write_cell(ws, r, 1, "  [t-statistic]", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTFOLIOS, 2):
            if p in data.index:
                t = data.loc[p, "t-stat (NW)"]
                write_cell(ws, r, i, f"({t:.3f})", FONT_BODY, ALIGN_C)
        r += 1

        # Alpha annualised
        write_cell(ws, r, 1, "α (% per year)", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTFOLIOS, 2):
            if p in data.index:
                pa = data.loc[p, "Alpha (% pa)"]
                write_cell(ws, r, i, f"{pa:.3f}", FONT_BODY, ALIGN_C)
        r += 1

        # Adj R²
        write_cell(ws, r, 1, "Adj. R²", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTFOLIOS, 2):
            if p in data.index:
                write_cell(ws, r, i, fmt(data.loc[p, "Adj R²"], 4), FONT_BODY, ALIGN_C)
        r += 1

        # Bottom border on last row of panel
        style_row(ws, r-1, 1, 7, border=Border(bottom=THIN))
        r += 2

    # Final bottom border
    style_row(ws, r-3, 1, 7, border=Border(bottom=THICK))

    # Notes
    r += 1
    ws.cell(r, 1, note_text).font = FONT_NOTE
    ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=7)


# ══════════════════════════════════════════════════════════════════════
# TABLE 5 — SUMMARY OF HYPOTHESES AND RESULTS
# ══════════════════════════════════════════════════════════════════════

def build_table5(wb):
    ws = wb.create_sheet("Table 5 Hypotheses")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14

    r = 1
    ws.cell(r,1,"Table 5").font = FONT_TITLE
    ws.cell(r,1).alignment = ALIGN_L
    r += 1
    ws.cell(r,1,"Summary of Hypotheses and Empirical Results").font = FONT_TITLE
    ws.cell(r,1).alignment = ALIGN_L
    r += 2

    hdrs = ["H","Hypothesis","Key Result","Evidence","Supported?"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(r, i, h)
        c.font = FONT_BOLD
        c.alignment = ALIGN_C if i > 1 else ALIGN_C
        c.border = Border(top=THICK, bottom=THIN)
    r += 1

    rows = [
        ("H1", "Alpha hypothesis",
         "Global EW portfolios earn positive and significant FF5 alpha",
         "All 5 quintiles: α > 0, t > 2.73 (p < 0.01). "
         "P5: α = 1.45% pm (17.5% pa, t = 4.93***). "
         "HML spread: α = 0.53% pm (t = 2.73***).",
         "Supported"),

        ("H2", "Monotonicity hypothesis",
         "FF5 alpha increases monotonically across ownership quintiles",
         "P1=0.924%, P2=1.135%, P3=1.378%, P4=1.090%, P5=1.454%. "
         "P4 dips below P3: not fully monotonic. "
         "However, P5 > P1 significantly (HML t = 2.73***).",
         "Partially"),

        ("H3", "Specification robustness",
         "Ownership alpha robust to FF6 (+ momentum) and value-weighting",
         "FF6 HML: α = 0.528% pm (t = 2.79***), vs FF5: 0.530%. "
         "Momentum factor (UMD) does not absorb the alpha.",
         "Supported"),

        ("H4", "Regional robustness",
         "Positive alpha holds in each regional subsample",
         "All P5 alphas positive and significant across all 4 regions. "
         "HML significant in North America (t = 1.68*). "
         "Asia-Pac HML positive but not significant (t = 1.63). "
         "EM HML positive but not significant (t = 0.94).",
         "Partially"),

        ("H5", "US-Europe contrast",
         "North America ownership alpha exceeds European alpha",
         "NA HML: +1.768% pm (21.2% pa, t = 1.68*). "
         "EU HML: −0.133% pm (−1.6% pa, t = −0.47). "
         "Difference is large and directionally consistent.",
         "Supported"),
    ]

    for hyp, hyp_name, hypothesis, result, support in rows:
        ws.cell(r,1,hyp).font = FONT_BOLD
        ws.cell(r,1).alignment = ALIGN_C
        ws.cell(r,2,hyp_name).font = FONT_BOLD
        ws.cell(r,2).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(r,3,hypothesis).font = FONT_BODY
        ws.cell(r,3).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(r,4,result).font = FONT_BODY
        ws.cell(r,4).alignment = Alignment(horizontal="left", wrap_text=True)
        col = support == "Supported" and "4CAF50" or (support == "Partially" and "FF9800" or "F44336")
        c = ws.cell(r,5,support)
        c.font = Font(name="Times New Roman", size=10, bold=True,
                      color=("1B5E20" if support == "Supported" else "E65100"))
        c.alignment = ALIGN_C
        ws.row_dimensions[r].height = 60
        r += 1

    style_row(ws, r-1, 1, 5, border=Border(bottom=THICK))

    r += 1
    note = ("Notes: Supported = hypothesis fully supported; Partially = partially supported. "
            "***p<0.01, **p<0.05, *p<0.10. All regressions use Newey-West (1987) "
            "heteroskedasticity- and autocorrelation-consistent standard errors with 6 lags. "
            "Sample: January 2015 to December 2025 (132 months), 1,285 globally listed firms "
            "with insider ownership > 20% and market capitalisation > $1 billion (Wood, 2025). "
            "Global regressions use Fama French Developed Markets five factor model. "
            "Regional regressions use regional factors from Fama and French (2012, 2015).")
    ws.cell(r,1,note).font = FONT_NOTE
    ws.cell(r,1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=5)


# ══════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════

def build_r1(wb):
    ws = wb.create_sheet("Table R1 Split Sample")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    for c in ["B","C","D","E","F","G"]:
        ws.column_dimensions[c].width = 13

    r = 1
    wc(ws,r,1,"Table R1", FONT_TITLE, ALIGN_L)
    r += 1
    wc(ws,r,1,"Split Sample Robustness: FF5 Alpha by Ownership Quintile", FONT_TITLE, ALIGN_L)
    r += 2

    col_labels = ["P1\n(Low)","P2","P3","P4","P5\n(High)","HML\n(P5−P1)"]
    PORTS = ["P1","P2","P3","P4","P5","HML"]

    for panel_label, period_label, data, n_months in [
        ("Panel A: Pre-COVID", "January 2015 to December 2019 (60 months)", res_pre,  60),
        ("Panel B: Post-COVID","January 2020 to December 2025 (72 months)", res_post, 72),
    ]:
        wc(ws,r,1,f"{panel_label}: {period_label}", FONT_PANEL, ALIGN_L)
        r += 1

        # Column header row
        wc(ws,r,1,"",FONT_BOLD,ALIGN_L,Border(top=THICK,bottom=THIN))
        for i, lbl in enumerate(col_labels, 2):
            c = wc(ws,r,i,lbl,FONT_BOLD,ALIGN_C,Border(top=THICK,bottom=THIN))
            c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        r += 1

        # Alpha
        wc(ws,r,1,"α (% per month)", FONT_BOLD, ALIGN_L)
        for i, p in enumerate(PORTS, 2):
            if p in data.index:
                a  = data.loc[p,"Alpha (% pm)"]
                pv = data.loc[p,"p-value"]
                wc(ws,r,i,f"{a:.4f}{stars(pv)}", FONT_BOLD, ALIGN_C)
        r += 1

        # T-stat
        wc(ws,r,1,"  [t-statistic]", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTS, 2):
            if p in data.index:
                wc(ws,r,i,f"({data.loc[p,'t-stat (NW)']:.3f})", FONT_BODY, ALIGN_C)
        r += 1

        # Annual alpha
        wc(ws,r,1,"α (% per year)", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTS, 2):
            if p in data.index:
                wc(ws,r,i,f"{data.loc[p,'Alpha (% pa)']:.3f}", FONT_BODY, ALIGN_C)
        r += 1

        # Adj R²
        wc(ws,r,1,"Adj. R²", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTS, 2):
            if p in data.index:
                wc(ws,r,i,f"{data.loc[p,'Adj R²']:.4f}", FONT_BODY, ALIGN_C)
        r += 1

        # N months
        wc(ws,r,1,"Observations (months)", FONT_BODY, ALIGN_L)
        for i, p in enumerate(PORTS, 2):
            if p in data.index:
                wc(ws,r,i,int(data.loc[p,"N months"]), FONT_BODY, ALIGN_C)

        # Separator
        for col in range(1, 8):
            ws.cell(r, col).border = Border(bottom=THIN)
        r += 3

    # Final border
    for col in range(1, 8):
        ws.cell(r-3, col).border = Border(bottom=THICK)

    # Note
    note = (
        "Notes: This table reports split sample Fama French five factor regression results "
        "to address the concern that static ownership sorting (2025 snapshot) may introduce "
        "look-ahead bias. If the ownership-alpha relationship is spuriously driven by "
        "end-of-sample ownership concentration, we would expect the alpha pattern to be "
        "substantially weaker or absent in the pre-2020 sub-period. "
        "The ownership quintile assignments are identical in both panels "
        "(sorted once using 2025 snapshot ownership). "
        "Global Fama French Developed Markets factors from French (2025) data library. "
        "Newey-West t-statistics with 6 lags. ***p<0.01, **p<0.05, *p<0.10."
    )
    ws.cell(r,1,note).font = FONT_NOTE
    ws.cell(r,1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=7)


# ── TABLE R2: Alternative sorts ───────────────────────────────────

def build_r2(wb):
    ws = wb.create_sheet("Table R2 Alt Sorts")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for c in ["B","C","D","E"]:
        ws.column_dimensions[c].width = 16

    r = 1
    wc(ws,r,1,"Table R2", FONT_TITLE, ALIGN_L)
    r += 1
    wc(ws,r,1,"Alternative Portfolio Sort Specifications: FF5 Alpha Robustness", FONT_TITLE, ALIGN_L)
    r += 2

    # ── Panel A: Tercile sort ──────────────────────────────────────
    wc(ws,r,1,f"Panel A: Tercile Sort  (cut-offs: {t1_break:.2f}% and {t2_break:.2f}%)", FONT_PANEL, ALIGN_L)
    r += 1

    hdrs_t3 = [
        ("T1 (Low)\n20–31%", "T1"),
        ("T2 (Mid)\n31–48%", "T2"),
        ("T3 (High)\n48–100%","T3"),
        ("HML\n(T3−T1)", "HML"),
    ]
    wc(ws,r,1,"",FONT_BOLD,ALIGN_L,Border(top=THICK,bottom=THIN))
    for i,(lbl,_) in enumerate(hdrs_t3, 2):
        c = wc(ws,r,i,lbl,FONT_BOLD,ALIGN_C,Border(top=THICK,bottom=THIN))
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    r += 1

    for row_lbl, fn in [
        ("N firms",         lambda p: cnt_t3.get(p,0)),
        ("α (% per month)", lambda p: f"{res_t3.loc[p,'Alpha (% pm)']:.4f}{stars(res_t3.loc[p,'p-value'])}" if p in res_t3.index else ""),
        ("[t-statistic]",   lambda p: f"({res_t3.loc[p,'t-stat (NW)']:.3f})" if p in res_t3.index else ""),
        ("α (% per year)",  lambda p: f"{res_t3.loc[p,'Alpha (% pa)']:.3f}" if p in res_t3.index else ""),
        ("Adj. R²",         lambda p: f"{res_t3.loc[p,'Adj R²']:.4f}" if p in res_t3.index else ""),
        ("Observations",    lambda p: int(res_t3.loc[p,'N months']) if p in res_t3.index else ""),
    ]:
        f = FONT_BOLD if "α (% per month)" in row_lbl else FONT_BODY
        wc(ws,r,1,("  " if "t-stat" in row_lbl else "")+row_lbl, f, ALIGN_L)
        for i,(_,p) in enumerate(hdrs_t3, 2):
            wc(ws,r,i,fn(p), f, ALIGN_C)
        r += 1

    for col in range(1, 6):
        ws.cell(r-1, col).border = Border(bottom=THIN)
    r += 2

    # ── Panel B: Decile sort ──────────────────────────────────────
    wc(ws,r,1,"Panel B: Decile Sort: Granular Ownership Alpha Pattern  (~128 firms per decile)", FONT_PANEL, ALIGN_L)
    r += 1

    DECILES   = [f"D{i}" for i in range(1,11)]
    dec_hdrs  = [(f"D{i}\n{d10_ranges[f'D{i}'][0]:.1f}–{d10_ranges[f'D{i}'][1]:.1f}%", f"D{i}") for i in range(1,11)]
    dec_hdrs += [("HML\n(D10−D1)", "HML")]
    n_dec_cols = len(dec_hdrs) + 1   # label col + data cols

    # Wider columns for decile table
    ws.column_dimensions["A"].width = 22
    for idx, letter in enumerate(["B","C","D","E","F","G","H","I","J","K","L"], 1):
        ws.column_dimensions[letter].width = 10

    wc(ws,r,1,"",FONT_BOLD,ALIGN_L,Border(top=THICK,bottom=THIN))
    for i,(lbl,_) in enumerate(dec_hdrs, 2):
        c = wc(ws,r,i,lbl,FONT_BOLD,ALIGN_C,Border(top=THICK,bottom=THIN))
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[r].height = 38
    r += 1

    for row_lbl, fn, bold in [
        ("α (% per month)", lambda p: f"{res_d10.loc[p,'Alpha (% pm)']:.4f}{stars(res_d10.loc[p,'p-value'])}" if p in res_d10.index else "", True),
        ("  [t-statistic]", lambda p: f"({res_d10.loc[p,'t-stat (NW)']:.3f})" if p in res_d10.index else "", False),
        ("α (% per year)",  lambda p: f"{res_d10.loc[p,'Alpha (% pa)']:.3f}" if p in res_d10.index else "", False),
        ("Adj. R²",         lambda p: f"{res_d10.loc[p,'Adj R²']:.4f}" if p in res_d10.index else "", False),
        ("Observations",    lambda p: str(int(res_d10.loc[p,'N months'])) if p in res_d10.index else "", False),
    ]:
        f = FONT_BOLD if bold else FONT_BODY
        wc(ws,r,1,row_lbl,f,ALIGN_L)
        for i,(_,p) in enumerate(dec_hdrs, 2):
            wc(ws,r,i,fn(p),f,ALIGN_C)
        r += 1

    for col in range(1, n_dec_cols+1):
        ws.cell(r-1,col).border = Border(bottom=THICK)
    r += 1

    note = (
        "Notes: This table tests whether the main result in Table 2 is sensitive to the "
        "specific ownership portfolio cut-off choices. "
        "Panel A uses equal-size tercile sorts. "
        "Panel B uses equal-size decile sorts (~128 firms each) to reveal the granular "
        "shape of the ownership-alpha relationship. The decile ranges are narrow at the "
        "bottom of the distribution (D1: 20–23%, D2: 23–26%) reflecting the "
        "clustering of firms just above the 20% threshold, and wider at the top "
        "(D10: 68–100%) reflecting the right-skewed ownership distribution. "
        "The decile progression shows whether the ownership-alpha relationship is smooth "
        "and continuous or concentrated in specific ownership ranges. "
        "HML is long D10 and short D1. All specifications use global Fama French "
        "Developed Markets five factor model with Newey-West t-statistics (6 lags), "
        "equal weighted returns, January 2015 to December 2025. "
        "***p<0.01, **p<0.05, *p<0.10."
    )
    ws.cell(r,1,note).font = FONT_NOTE
    ws.cell(r,1).alignment = Alignment(horizontal="left", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=12)




wb = Workbook()
wb.remove(wb.active)   # remove default empty sheet

# Table 1 — Sample characteristics
build_table1(wb)

# Table 2 — Global FF5 (main result)
build_regression_table(
    wb,
    sheet_name = "Table 2 FF5 Global",
    title      = "Table 2",
    subtitle   = ("Fama French Five Factor Alpha by Insider Ownership Quintile: "
                  "Global Equal Weighted Portfolios"),
    data       = ff5,
    factor_cols= ["β Mkt-RF","β SMB","β HML","β RMW","β CMA"],
    note_text  = (
        "Notes: This table reports Fama French five factor time-series regression results "
        "for ownership-sorted portfolios following Fahlenbrach (2009, Table 6). "
        "Firms are sorted once into quintiles based on aggregate insider ownership as of 2025. "
        "P1 (Low) contains firms with 20.00–26.46% ownership; P5 (High) contains firms with "
        "56.87–99.81% ownership. The HML portfolio is long P5 and short P1. "
        "Returns are equal weighted and winsorised at the 1st and 99th percentile. "
        "Regression specification: R_p,t − RF_t = α_p + β₁MKT_t + β₂SMB_t + β₃HML_t "
        "+ β₄RMW_t + β₅CMA_t + ε_p,t. "
        "Global Fama French Developed Markets factors from French (2025) data library. "
        "t-statistics in brackets are Newey-West (1987) corrected with 6 lags. "
        "Sample: January 2015 to December 2025 (132 months). "
        "***p<0.01, **p<0.05, *p<0.10."
    )
)

# Table 3 — Global FF6 (H3 robustness)
build_regression_table(
    wb,
    sheet_name = "Table 3 FF6 Robustness",
    title      = "Table 3",
    subtitle   = ("Fama French Six Factor Alpha by Insider Ownership Quintile: "
                  "Robustness Check Including Momentum Factor"),
    data       = ff6,
    factor_cols= ["β Mkt-RF","β SMB","β HML","β RMW","β CMA","β UMD"],
    note_text  = (
        "Notes: This table extends Table 2 by adding the Developed Momentum Factor (WML/UMD) "
        "to the five factor specification, following Carhart (1997). "
        "The six factor model tests whether the ownership alpha in Table 2 is subsumed by "
        "momentum exposures. Comparing alphas across Tables 2 and 3 shows that the inclusion "
        "of UMD leaves all portfolio alphas virtually unchanged: the global HML alpha is "
        "0.528% per month in the FF6 model versus 0.530% in the FF5 model, confirming "
        "robustness to momentum (H3). All other specifications identical to Table 2. "
        "***p<0.01, **p<0.05, *p<0.10."
    )
)

# Table 4 — Regional results (H4 and H5)
build_table4(
    wb,
    reg_data  = reg_sheets,
    title     = "Table 4",
    subtitle  = ("Regional Fama French Five Factor Alpha by Insider Ownership Quintile "
                 "(H4 and H5)"),
    note_text = (
        "Notes: This table reports FF5 regression results separately for four regional "
        "subsamples, testing Hypotheses H4 (regional robustness) and H5 (US-Europe contrast). "
        "Panel A uses Fama French North American five factor returns; Panel B uses European "
        "five factor returns; Panel C uses Asia Pacific ex Japan five factor returns; "
        "Panel D uses global Developed Markets factors as a proxy (no EM ex-Asia regional "
        "factor set is published by French). "
        "Panel A (N=141): HML alpha = 1.768% per month (21.2% per year), significant at "
        "the 10% level (t = 1.681). Panel B (N=109): HML alpha = −0.133% per month "
        "(−1.6% per year), not significant. This confirms H5: the North American ownership "
        "alpha substantially exceeds the European alpha, consistent with the scarcity "
        "of high-ownership firms in US capital markets documented by Wood (2025). "
        "t-statistics are Newey-West corrected (6 lags). ***p<0.01, **p<0.05, *p<0.10."
    )
)

# Table 5 — Hypothesis summary
build_table5(wb)

# Save
out = "thesis_tables.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print("Sheets created:")


# Add robustness tables R1 and R2 to the workbook
build_r1(wb)
build_r2(wb)
build_regression_table(
    wb,
    sheet_name  = "Table R3 FF5 VW",
    title       = "Table R3",
    subtitle    = ("Fama French Five Factor Alpha by Insider Ownership Quintile: "
                   "Value Weighted Portfolios (Robustness)"),
    data        = vw,
    factor_cols = ["β Mkt-RF","β SMB","β HML","β RMW","β CMA"],
    note_text   = (
        "Notes: This table replicates Table 2 using value weighted portfolio returns "
        "as a robustness check. Within each ownership quintile, firms are weighted by "
        "their lagged market capitalisation (end of previous month) from FactSet "
        "(FREF_MARKET_VALUE_COMPANY). Value-weighting reduces the influence of smaller "
        "firms within each portfolio and tests whether the equal weighted alpha in "
        "Table 2 is driven by a size effect rather than an ownership effect. "
        "All other specifications are identical to Table 2: global Fama French "
        "Developed Markets five factor model, Newey-West t-statistics (6 lags), "
        "January 2015 to December 2025 (132 months). "
        "***p<0.01, **p<0.05, *p<0.10."
    )
)
wb.save("thesis_tables.xlsx")
print("\nSaved: thesis_tables.xlsx  (Tables 1–5, R1, R2)")
print("Sheets:", [s for s in wb.sheetnames])


# ══════════════════════════════════════════════════════════════════════
# ROBUSTNESS TABLES — R1 (split-sample) and R2 (alternative sorts)
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# FIGURES — Academic graphs Figures 1–6
# ══════════════════════════════════════════════════════════════════════

plt.rcParams.update({
    # Font
    "font.family":        "serif",
    "font.serif":         ["Times New Roman", "DejaVu Serif", "serif"],
    "font.size":          11,
    "axes.titlesize":     12,
    "axes.labelsize":     11,
    "xtick.labelsize":    10,
    "ytick.labelsize":    10,
    "legend.fontsize":    10,
    # Lines and axes
    "axes.linewidth":     0.8,
    "axes.edgecolor":     "black",
    "axes.spines.top":    False,
    "axes.spines.right":  False,
    "xtick.direction":    "out",
    "ytick.direction":    "out",
    "xtick.major.size":   4,
    "ytick.major.size":   4,
    # Grid
    "axes.grid":          True,
    "grid.color":         "#e0e0e0",
    "grid.linewidth":     0.5,
    # Figure
    "figure.dpi":         150,
    "savefig.dpi":        300,
    "savefig.bbox":       "tight",
    "savefig.pad_inches": 0.05,
    # Legend
    "legend.frameon":     True,
    "legend.framealpha":  1.0,
    "legend.edgecolor":   "black",
    "legend.fancybox":    False,
})

# Colour palette — minimal academic style

PORTFOLIOS   = ["P1", "P2", "P3", "P4", "P5", "HML"]
REGIONS      = ["North America", "Europe", "Asia-Pacific", "EM ex-Asia"]




def save_fig(fig, name):
    """Save figure as both PDF and PNG."""
    fig.savefig(f"{name}.pdf")
    fig.savefig(f"{name}.png")
    print(f"  Saved: {name}.pdf  /  {name}.png")


# ══════════════════════════════════════════════════════════════════════
# LOAD RESULTS
# ══════════════════════════════════════════════════════════════════════

# (result files already loaded above)


# ══════════════════════════════════════════════════════════════════════
# FIGURE 1 — FF5 ALPHA BY OWNERSHIP QUINTILE  (H1 and H2)
# ══════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════
# FIGURE 6 — DECILE ALPHA CHART  (granular ownership-return pattern)
# Shows FF5 alpha for D1–D10 with ownership ranges on x-axis.
# Reads from results_alternative_sorts_FF5.xlsx (Decile_Sort sheet).
# This is the key visual argument for the granular ownership premium.
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# FIGURE 6 — DECILE ALPHA CHART
# Clean academic style following Wood (2025, IRFA) figure conventions:
# single bar color, no gradient, no colorbar, minimal decoration
# ══════════════════════════════════════════════════════════════════════

print("Building Figure 6: Decile alpha chart...")

try:
    dec_res = res_d10   # already loaded at top of script

    dec_range_labels = {
        "D1":"20.0–23.4%", "D2":"23.4–26.4%", "D3":"26.5–29.6%",
        "D4":"29.6–34.0%", "D5":"34.0–38.8%", "D6":"38.8–43.6%",
        "D7":"43.7–49.9%", "D8":"49.9–56.7%", "D9":"56.8–68.3%",
        "D10":"68.4–99.8%",
    }
    DECILES = [f"D{i}" for i in range(1, 11)]

    alphas  = [dec_res.loc[d, "Alpha (% pm)"] if d in dec_res.index else np.nan for d in DECILES]
    pvalues = [dec_res.loc[d, "p-value"]      if d in dec_res.index else 1.0    for d in DECILES]
    tstats  = [dec_res.loc[d, "t-stat (NW)"]  if d in dec_res.index else np.nan for d in DECILES]
    hml_a   = dec_res.loc["HML","Alpha (% pm)"] if "HML" in dec_res.index else None
    hml_t   = dec_res.loc["HML","t-stat (NW)"]  if "HML" in dec_res.index else None
    hml_pv  = dec_res.loc["HML","p-value"]       if "HML" in dec_res.index else None

    # Two-line x-axis labels: decile number + ownership range
    xlabels = [f"D{i}\n({dec_range_labels[f'D{i}']})" for i in range(1, 11)]

    # ── Figure layout ──────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(9, 4.5))

    x    = np.arange(len(DECILES))
    BAR_COLOR = "#2c2c2c"   # single dark grey: clean, no gradient

    bars = ax.bar(x, alphas, width=0.55,
                  color=BAR_COLOR, edgecolor="black", linewidth=0.5, zorder=3)

    # ── Significance stars above each bar ─────────────────────────
    for bar, pv, a in zip(bars, pvalues, alphas):
        if not np.isnan(a):
            s = stars(pv)
            if s:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        a + 0.018, s,
                        ha="center", va="bottom",
                        fontsize=9, fontweight="bold", color="black")

    # ── t-statistics below each bar label (not inside) ────────────
    for i, (tstat, a) in enumerate(zip(tstats, alphas)):
        if not np.isnan(tstat):
            ax.text(x[i], -0.09, f"({tstat:.2f})",
                    ha="center", va="top",
                    fontsize=7, color="#333333")

    # ── Linear trend line (thin, dashed, subtle) ──────────────────
    z    = np.polyfit(x, alphas, 1)
    poly = np.poly1d(z)
    xfit = np.linspace(-0.4, len(DECILES) - 0.6, 300)
    ax.plot(xfit, poly(xfit), color="black", linewidth=1.0,
            linestyle="--", zorder=4, alpha=0.55, label="Linear trend")

    # ── Zero line ─────────────────────────────────────────────────
    ax.axhline(0, color="black", linewidth=0.7, zorder=4)

    # ── Axes and labels ───────────────────────────────────────────
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, fontsize=8)
    ax.set_ylabel("FF5 Alpha (% per month)", fontsize=10, labelpad=8)
    ax.set_xlabel("Insider Ownership Decile", fontsize=10, labelpad=28)
    ax.set_ylim(-0.15, max(alphas) + 0.22)

    # Spine cleanup — top and right off (Wood style)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.yaxis.grid(True, color="#dddddd", linewidth=0.5, zorder=0)
    ax.set_axisbelow(True)

    # ── Title ─────────────────────────────────────────────────────
    ax.set_title(
        "Figure 6.  FF5 Alpha by Insider Ownership Decile\n"
        "Equal Weighted Portfolios, January 2015 to December 2025",
        fontsize=10, loc="left", pad=8
    )

    # ── Legend ────────────────────────────────────────────────────
    ax.legend(fontsize=8, frameon=False, loc="upper left")

    # ── HML note in top right corner ──────────────────────────────
    if hml_a is not None:
        hml_s = stars(hml_pv)
        ax.text(0.98, 0.97,
                f"HML (D10 − D1): {hml_a:.3f}%{hml_s}  (t = {hml_t:.2f})",
                transform=ax.transAxes,
                ha="right", va="top", fontsize=8.5, style="italic",
                bbox=dict(boxstyle="square,pad=0.3", facecolor="white",
                          edgecolor="#aaaaaa", linewidth=0.6))

    # ── Note below figure ─────────────────────────────────────────
    fig.text(
        0.01, -0.08,
        "Note: This figure shows the equal weighted Fama French five factor alpha "
        "for ten ownership-sorted decile portfolios. Each decile contains approximately "
        "128 firms. Ownership ranges are shown on the x-axis. "
        "Newey-West (1987) t-statistics with 6 lags in parentheses below each bar. "
        "The dashed line shows the linear trend. "
        "***p < 0.01,  **p < 0.05,  *p < 0.10.",
        fontsize=8, style="italic", ha="left", va="top",
        wrap=True
    )

    plt.tight_layout()
    save_fig(fig, "figure6_decile_alphas")
    plt.close()
    print("  Saved: figure6_decile_alphas.pdf / .png")

except FileNotFoundError:
    print("  results_alternative_sorts_FF5.xlsx not found: run thesis_analysis.py first")


print("\n" + "="*60)
print("DONE: outputs saved:")
print("  thesis_tables.xlsx         : all 8 academic tables (1–5, R1, R2, R3)")
print("  figure6_decile_alphas.pdf  : decile alpha chart (for thesis)")
print("  figure6_decile_alphas.png  : decile alpha chart (preview)")
print("="*60)
